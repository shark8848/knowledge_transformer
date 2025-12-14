"""Convert Excel workbooks (xls/xlsx) into Markdown tables per sheet."""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook
from tabulate import tabulate

from ..base import ConversionInput, ConversionPlugin, ConversionResult
from ..registry import REGISTRY


def _sheet_to_markdown(sheet) -> str:
    rows: List[List[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in row])
    if not rows:
        return f"### {sheet.title}\n\n(空工作表)\n"

    table_md = tabulate(rows, headers="firstrow", tablefmt="github")
    return f"### {sheet.title}\n\n{table_md}\n"


class ExcelToMarkdownPlugin(ConversionPlugin):
    slug = "excel-to-md"
    source_format = "xlsx"
    target_format = "md"

    def convert(self, payload: ConversionInput) -> ConversionResult:
        if not payload.input_path:
            raise ValueError("Conversion requires local input_path for Excel files")

        input_path = Path(payload.input_path)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")

        wb = load_workbook(filename=str(input_path), data_only=True, read_only=True)
        parts: List[str] = []
        for sheet in wb.worksheets:
            parts.append(_sheet_to_markdown(sheet))
        wb.close()

        if not parts:
            md = "(空工作簿)"
        else:
            md = "\n\n".join(parts)

        output_path = input_path.with_suffix(".md")
        output_path.write_text(md, encoding="utf-8")

        metadata = {"note": "Converted Excel to Markdown"}
        return ConversionResult(output_path=output_path, metadata=metadata)


class ExcelLegacyToMarkdownPlugin(ExcelToMarkdownPlugin):
    slug = "xls-to-md"
    source_format = "xls"


REGISTRY.register(ExcelToMarkdownPlugin)
REGISTRY.register(ExcelLegacyToMarkdownPlugin)
